import os, shutil, time
from pathlib import Path


CurrentDirectory = os.path.dirname(__file__)
dirInput = os.path.join(CurrentDirectory, "input1")
rutTest = ""
# rutTest = "76487348-3"

def getDescItem():

    import lxml
    import xml.etree.ElementTree as ET
    from lxml import etree
    from lxml import etree, objectify
    proveedores = []
    for elem in os.walk(dirInput):
        if elem[0] == dirInput:
            for arch in elem[2]:
                dirArch = os.path.join(dirInput, arch)
                treeArch = ET.parse(dirArch)
                try:
                    prov = treeArch.findall(".//{*}Encabezado/{*}Emisor/{*}RUTEmisor")[0].text
                    descItem = treeArch.findall(".//{*}Detalle/{*}DscItem")[0].text
                    detalles = treeArch.findall(".//{*}Documento/{*}Detalle")
                    if(len(detalles) > 0):
                        print("dirArch", dirArch)
                        proveedores.append(prov)

                except:
                    pass

def getEstadisticaInput():

    def llenarPlantilla(tupEscReporte):

        def getOrgCatalogo(RutOrg):
            dirCatalogoOrg = os.path.join(CurrentDirectory, "catalogos", "organizaciones.xlsx")
            wbCatalogoOrg = op.load_workbook(filename=dirCatalogoOrg)
            wsCatalogoOrg = wbCatalogoOrg.active
            lastEmptyCatalogoOrg = len(list(wsCatalogoOrg))
            k = 1
            nombreOrg = ""
            while k < lastEmptyCatalogoOrg:
                if(wsCatalogoOrg.cell(k, 1).value == RutOrg):
                    nombreOrg = wsCatalogoOrg.cell(k, 2).value
                k+=1
                wbCatalogoOrg.close()
            return nombreOrg

        import openpyxl as op
        dirPlantilla = os.path.join(CurrentDirectory, "catalogos", "plantillaEstadisticasInput.xlsx")
        nuevaDirPlantilla = os.path.join(CurrentDirectory, "plantillaEstadisticasInput.xlsx")
        if os.path.isfile(nuevaDirPlantilla):
            os.remove(nuevaDirPlantilla)
        shutil.copy(dirPlantilla, CurrentDirectory)
        wb = op.load_workbook(filename=nuevaDirPlantilla)
        ws = wb.active
        ws.cell(1, 1).value = "Estadísticas de los proveedores"
        ws.cell(2, 1).value = "Número de Identificacion"
        ws.cell(2, 2).value = "Nombre de la Organización"
        ws.cell(2, 3).value = "total xml"
        ws.cell(2, 4).value = "xml con shipment"
        ws.cell(2, 5).value = "xml sin shipment"
        k = 3

        for tup in tupEscReporte:
            rutOrg = ""
            rutOrg = getOrgCatalogo(tup[0])
            ws.cell(k, 1).value = tup[0]
            ws.cell(k, 2).value = rutOrg
            ws.cell(k, 3).value = tup[1]
            ws.cell(k, 4).value = tup[2]
            ws.cell(k, 5).value = tup[3]
            k+=1

        wb.save(filename=nuevaDirPlantilla)
        wb.close()

    dirInput = os.path.join(CurrentDirectory, "input")

    tupEscReporte = []
    for elem in os.walk(dirInput):
        if(elem[0] == dirInput):
            for subArch in elem[2]:
                dirSubArch = os.path.join(dirInput, subArch)
                numIden = subArch.split("_")[0]
                f = open(dirSubArch, "r")
                SN = ""
                linea = ""
                try:
                    for line in f.readlines():
                        if("S0" in line):
                            palabras = line.split("S")

                            for pal in palabras:
                                if(pal[0:7].isnumeric() and len(pal[0:8]) == 8):
                                    SN = "S"+pal[0:8]
                                    if(linea == ""):
                                        linea = line

                except:
                    pass

                if len(SN) == 9 and SN[0] == "S" and (SN[1:9].isnumeric()):
                    encontrado = False
                    for tup in tupEscReporte:
                        if(numIden == tup[0]):
                            tup[1]+=1
                            tup[2]+=1
                            encontrado = True
                    if(encontrado == False):
                        tupEscReporte.append([numIden, 1, 1, 0])

                elif(SN == ""):
                    for tup in tupEscReporte:
                        if(numIden == tup[0]):
                            tup[1]+=1
                            tup[3]+=1
                            encontrado = True

                    if(encontrado == False):
                        tupEscReporte.append([numIden, 1, 0, 1])
    # llenarPlantilla(tupEscReporte)

def asegurarEstructuraTransacciones():

    import lxml
    import xml.etree.ElementTree as ET
    from lxml import etree
    from lxml import etree, objectify

    def validarEstructuraIdDoc(estructuraTree):
        estBase = False
        try:
            estructuraTree.findall(".//{*}Encabezado/{*}IdDoc/{*}TipoDTE")[0].text
            estructuraTree.findall(".//{*}Encabezado/{*}IdDoc/{*}Folio")[0].text
            # estructuraTree.findall(".//{*}Encabezado/{*}IdDoc/{*}FchEmis")[0].text
            # estructuraTree.findall(".//{*}Encabezado/{*}IdDoc/{*}FmaPago")[0].text
            estBase = True
        except:
            pass
        return estBase

    def validarTED(estructuraTree):
        estBase = False
        try:
            RE = estructuraTree.findall(".//{*}TED/{*}DD/{*}RE")[0].text
            TD = estructuraTree.findall(".//{*}TED/{*}DD/{*}TD")[0].text
            F = estructuraTree.findall(".//{*}TED/{*}DD/{*}F")[0].text
            FE = estructuraTree.findall(".//{*}TED/{*}DD/{*}FE")[0].text
            RR = estructuraTree.findall(".//{*}TED/{*}DD/{*}RR")[0].text
            RSR = estructuraTree.findall(".//{*}TED/{*}DD/{*}RSR")[0].text
            MNT = estructuraTree.findall(".//{*}TED/{*}DD/{*}MNT")[0].text
            IT1 = estructuraTree.findall(".//{*}TED/{*}DD/{*}IT1")[0].text
            # RE = estructuraTree.findall(".//{*}TED/{*}DD/{*}CAF/{*}DA/{*}RE")[0].text
            # RS = estructuraTree.findall(".//{*}TED/{*}DD/{*}CAF/{*}DA/{*}RS")[0].text
            # TD = estructuraTree.findall(".//{*}TED/{*}DD/{*}CAF/{*}DA/{*}TD")[0].text
            # D = estructuraTree.findall(".//{*}TED/{*}DD/{*}CAF/{*}DA/{*}RNG/{*}D")[0].text
            # H = estructuraTree.findall(".//{*}TED/{*}DD/{*}CAF/{*}DA/{*}RNG/{*}H")[0].text
            # FA = estructuraTree.findall(".//{*}TED/{*}DD/{*}CAF/{*}DA/{*}FA")[0].text
            # M = estructuraTree.findall(".//{*}TED/{*}DD/{*}CAF/{*}DA/{*}RSAPK/{*}M")[0].text
            # E = estructuraTree.findall(".//{*}TED/{*}DD/{*}CAF/{*}DA/{*}RNG/{*}E")[0].text
            # IDK = estructuraTree.findall(".//{*}TED/{*}DD/{*}CAF/{*}DA/{*}IDK")[0].text
            # TSTED = estructuraTree.findall(".//{*}TED/{*}DD/{*}TSTED")[0].text

            estBase = True
        except:
            pass
        return estBase

    dirInput = os.path.join(CurrentDirectory, "input1")
    totalArch = 0
    totalDTE = 0
    estructuraIdDoc = 0
    estructuraTED = 0

    for elem in os.walk(dirInput):
        if (elem[0] == dirInput):
            totalArch = len(elem[2])
            for subArch in elem[2]:
                dirSubArch = os.path.join(dirInput, subArch)
                estructuraTree = ET.parse(dirSubArch)
                try:
                    DTE = estructuraTree.findall(".//{*}DTE")
                    totalDTE+=1
                except:
                    pass

                if(validarEstructuraIdDoc(estructuraTree)):
                    estructuraIdDoc+=1

                if(validarTED(estructuraTree)):
                    estructuraTED+=1

    print("totalArch", totalArch)
    print("totalDTE", totalDTE)
    print("idDoc(TipoDTE, Folio, FchEmis, FmaPago)", estructuraIdDoc)
    print("idDoc(RE, TD, F, FE, RR, RSR, MNT, IT1)", estructuraTED)

def actualizarCatalogoOrgs():

    import openpyxl as op
    import lxml
    import xml.etree.ElementTree as ET
    from lxml import etree
    from lxml import etree, objectify

    def getInfoOrgCW(RutOrg):

        codeOrg = ""
        nombreOrg = ""

        from datetime import datetime

        import openpyxl as op
        import lxml
        import xml.etree.ElementTree as ET
        from lxml import etree
        from lxml import etree, objectify

        import logging
        import glob
        import sys
        import requests
        from requests.auth import HTTPBasicAuth
        from unidecode import unidecode as ud

        orgLocation = os.path.join(CurrentDirectory, "catalogos", "organization")
        if os.path.isdir(orgLocation):
            shutil.rmtree(orgLocation)
        os.mkdir(orgLocation)

        def my_custom_logger(logger_name, level=logging.DEBUG):

            """
            Method to return a custom logger with the given name and level
            """
            logger = logging.getLogger(logger_name)
            logger.setLevel(level)
            format_string = ("%(asctime)s — %(name)s — %(levelname)s — %(funcName)s:"
                             "%(lineno)d — %(message)s")
            log_format = logging.Formatter(format_string)
            # Creating and adding the console handler
            console_handler = logging.StreamHandler(sys.stdout)
            console_handler.setFormatter(log_format)
            logger.addHandler(console_handler)
            # Creating and adding the file handler
            file_handler = logging.FileHandler(logger_name, mode='a')
            file_handler.setFormatter(log_format)
            logger.addHandler(file_handler)
            return logger

        def ExtraerXML(InputFolderXML, OutputFolderXML):

            url = 'https://a5lprdservices.wisegrid.net/eAdaptor'

            user = "A5L.data.entry"

            dirClave = os.path.join(os.path.dirname(__file__), "clave.txt")
            password = open(dirClave).read()

            dirname = os.path.dirname(__file__)
            responseFiles = []
            for file in glob.glob(os.path.join(InputFolderXML, "*.xml")):
                try:
                    filename = os.path.basename(file)
                    data = open(file, encoding='utf-8').read()

                    response = requests.post(url, data=data.encode('utf-8'), auth=(user, password))
                    ResponseFile = open(os.path.join(OutputFolderXML, "Response_" + filename), "w+")
                    ResponseFile.write(ud(str(response.text))[4:])
                    ResponseFile.close()
                    i = 1

                    newfilename = filename
                    while os.path.exists(os.path.join(OutputFolderXML, newfilename)):
                        newfilename = filename[:-4] + "(" + str(i) + ").xml"
                        i += 1

                    direccionFila = os.path.join(InputFolderXML, filename)
                    os.remove(direccionFila)

                except Exception as e:
                    logger = my_custom_logger(f"Response/error_log_{filename[:-4]}.log")
                    logger.debug(str(e))
            return responseFiles

        NativeOrganizationRequest = ET.Element("Native", xmlns="http://www.cargowise.com/Schemas/Native")
        BodyOrganizationRequest = ET.SubElement(NativeOrganizationRequest, "Body")

        Organization = ET.SubElement(BodyOrganizationRequest, "Organization")
        CriteriaGroupOrg = ET.SubElement(Organization, "CriteriaGroup", Type="Partial")
        CriteriaEntity = ET.SubElement(CriteriaGroupOrg, "Criteria", Entity="OrgHeader.OrgCusCode", FieldName="CustomsRegNo").text = RutOrg

        tree = ET.ElementTree(NativeOrganizationRequest)
        filename = os.path.join(orgLocation, datetime.today().strftime("%Y %m %d_%H-%M %f") + '_OrganizationRequest_' + str(RutOrg) + '.xml')
        tree.write(filename, encoding="utf-8", xml_declaration=False, short_empty_elements=False)
        prettytree = lxml.etree.parse(filename)

        with open(filename, "wb") as f:
            f.write(etree.tostring(prettytree, pretty_print=True, xml_declaration=False, encoding="UTF-8"))

        responseFiles = ExtraerXML(orgLocation, orgLocation)

        for elem in os.walk(orgLocation):
            if elem[0] == orgLocation:
                for arch in elem[2]:
                    if ("Response_" in arch and "_OrganizationRequest_" in arch and ".xml" in arch):
                        dirArch = os.path.join(orgLocation, arch)
                        treeOrg = ET.parse(dirArch)
                        dataSourceOrgCusCodeCollection = treeOrg.findall('.//{*}Organization/{*}OrgHeader/{*}OrgCusCodeCollection')
                        orgCorrecta = False
                        for orgCusCode in dataSourceOrgCusCodeCollection:
                            num = orgCusCode.findall(".//{*}CustomsRegNo")[0].text
                            if (num == RutOrg):
                                orgCorrecta = True

                        if (orgCorrecta == True):
                            codeOrg = treeOrg.findall('.//{*}Organization/{*}OrgHeader/{*}Code')[0].text
                            nombreOrg = treeOrg.findall('.//{*}Organization/{*}OrgHeader/{*}FullName')[0].text
                        os.remove(dirArch)

        return nombreOrg, codeOrg

    dirCatalogoOrg = os.path.join(CurrentDirectory, "catalogos", "organizaciones.xlsx")
    if os.path.isfile(dirCatalogoOrg) == False:
        wb = op.Workbook()
        Sheet_name = wb.sheetnames
        wb.save(filename=dirCatalogoOrg)

    if os.path.isfile(dirCatalogoOrg):
        listaRegNum = []
        for elem in os.walk(dirInput):
            if (elem[0] == dirInput):
                for subArch in elem[2]:
                    dirSubArch = os.path.join(dirInput, subArch)
                    treeInput = ET.parse(dirSubArch)
                    rutOrg = treeInput.findall('.//{*}Encabezado/{*}Emisor/{*}RUTEmisor')[0].text
                    if(rutOrg not in listaRegNum):
                        listaRegNum.append(rutOrg)

        wbCatalogoOrg = op.load_workbook(filename=dirCatalogoOrg)
        wsCatalogoOrg = wbCatalogoOrg.active
        for rut in listaRegNum:

            lastEmptyCatalogoOrg = len(list(wsCatalogoOrg))
            k = 1
            encontrado = False

            while k < lastEmptyCatalogoOrg:
                if(wsCatalogoOrg.cell(k, 1).value == rut):
                    encontrado = True
                    if(wsCatalogoOrg.cell(k, 2).value == None or wsCatalogoOrg.cell(k, 3).value == None):
                        wsCatalogoOrg.cell(k, 2).value, wsCatalogoOrg.cell(k, 3).value = getInfoOrgCW(rut)
                k+=1
            if(encontrado == False):
                wsCatalogoOrg.cell(lastEmptyCatalogoOrg+1, 2).value, wsCatalogoOrg.cell(lastEmptyCatalogoOrg+1, 3).value = getInfoOrgCW(rut)
                wsCatalogoOrg.cell(lastEmptyCatalogoOrg + 1, 1).value = rut

        wbCatalogoOrg.save(filename=dirCatalogoOrg)
        wbCatalogoOrg.close()

def generarEstructuraIteracion():

    import lxml
    import xml.etree.ElementTree as ET
    from lxml import etree
    from lxml import etree, objectify

    print("----------------------------------------------generarEstructuraProcesar------------------------------------------------")
    def evaluarSubArch(dirSubArch):

        import lxml
        import xml.etree.ElementTree as ET
        from lxml import etree
        from lxml import etree, objectify

        SN = ""
        sirve = False
        transaccionTree = ET.parse(dirSubArch)
        try:
            SN = transaccionTree.findall(".//{*}Referencia/{*}FolioRef")[0].text
        except:
            pass

        # if SN == "":
        #     try:
        #         f = open(dirSubArch, 'r')
        #
        #         for line in f.readlines():
        #             if ("S" in line):
        #
        #                 palabras = line.split("S")
        #
        #                 for pal in palabras:
        #                     pal = pal.strip().replace(" ", "").replace("\n", "")
        #
        #                     if (pal[0:8].isnumeric() and len(pal[0:8]) == 8):
        #                         SN = "S" + pal[0:8]
        #
        #         if len(SN) == 9 and SN[0] == "S" and (SN[1:9].isnumeric()):
        #             sirve = True
        #     except:
        #         pass

        if len(SN) == 9 and SN[0] == "S" and (SN[1:9].isnumeric()):
            sirve = True

        return sirve

    def getOrgCatalogo(RutOrg):

        def generalizarString(stringRec):

            import re
            from unicodedata import normalize
            stringReturn = ""
            if (stringRec != None):
                stringRec = re.sub(r"([^n\u0300-\u036f]|n(?!\u0303(?![\u0300-\u036f])))[\u0300-\u036f]+", r"\1", normalize("NFD", stringRec), 0, re.I)
                stringRec = normalize('NFC', stringRec)

                stringReturn = stringRec
                charRemove = []
                for char in stringRec:
                    if (char.isalnum() == False and char != " " and char != "-"):
                        charRemove.append(char)
                for char in charRemove:
                    stringReturn = stringReturn.replace(char, "")

            return str(stringReturn.upper())

        import openpyxl as op

        dirCatalogoOrg = os.path.join(CurrentDirectory, "catalogos", "organizaciones.xlsx")
        wbCatalogoOrg = op.load_workbook(filename=dirCatalogoOrg)
        wsCatalogoOrg = wbCatalogoOrg.active
        lastEmptyCatalogoOrg = len(list(wsCatalogoOrg))
        k = 1
        nombreOrg = "No Identificada"
        while k < lastEmptyCatalogoOrg:
            if(wsCatalogoOrg.cell(k, 1).value == RutOrg):
                nombreOrg = wsCatalogoOrg.cell(k, 3).value
            k+=1
        wbCatalogoOrg.save(filename=dirCatalogoOrg)
        wbCatalogoOrg.close()

        if(nombreOrg == None or nombreOrg == ""):
            nombreOrg = "No Identificada"
        else:
            nombreOrg = generalizarString(nombreOrg)

        return nombreOrg

    from datetime import datetime
    dirIter = os.path.join(CurrentDirectory, "procesando", "iteracion_"+datetime.today().strftime("%Y-%m-%d"))
    if(os.path.isdir(dirIter)):
        shutil.rmtree(dirIter)
    os.mkdir(dirIter)

    for elem in os.walk(dirInput):
        if (elem[0] == dirInput):
            for subArch in elem[2]:
                if(rutTest in subArch):
                    dirSubArch = os.path.join(dirInput, subArch)
                    treeInput = ET.parse(dirSubArch)
                    rutOrg = treeInput.findall('.//{*}Encabezado/{*}Emisor/{*}RUTEmisor')[0].text
                    idenOrg = getOrgCatalogo(rutOrg)+"_"+rutOrg
                    dirProcesarOrganization = os.path.join(dirIter, idenOrg)
                    dirProcesarOrganizationSinSN = os.path.join(dirIter, idenOrg, "Inaceptables")
                    dirProcesarOrganizationTransacciones = os.path.join(dirIter, idenOrg, "Procesar")

                    if os.path.isdir(dirProcesarOrganization) == False:
                        os.mkdir(dirProcesarOrganization)
                        os.mkdir(dirProcesarOrganizationSinSN)
                        os.mkdir(dirProcesarOrganizationTransacciones)

                    sirve = evaluarSubArch(dirSubArch)
                    if(sirve):
                        shutil.copy(dirSubArch, dirProcesarOrganizationTransacciones)

                    else:
                        shutil.copy(dirSubArch, dirProcesarOrganizationSinSN)
    return dirIter

def procesarEstructuraIteracion(dirIter):

    print("---------------------------------------------procesarEstructuraIteracion-----------------------------------------------")

    def getCarpetasTransacciones(dirIter):

        def generarCarpetasTransaccion(carpetaOrg):

            import lxml
            import xml.etree.ElementTree as ET
            from lxml import etree
            from lxml import etree, objectify

            dirTransaccionesOrg = os.path.join(carpetaOrg, "Procesar")
            dirTransaccionesOrgsinSN = os.path.join(carpetaOrg, "Inaceptables")
            tuplasProcesar = []
            carpetasProcesar = []
            for elem in os.walk(dirTransaccionesOrg):
                if (elem[0] == dirTransaccionesOrg):
                    for arch in elem[2]:
                        pathTransaccion = os.path.join(dirTransaccionesOrg, arch)

                        carpetaTransaccion = ""
                        nuevaTransaccion = ""

                        transaccionTree = ET.parse(pathTransaccion)
                        IdenOrg = ""
                        numTransaccion = ""
                        SN = ""

                        try:
                            IdenOrg = transaccionTree.findall(".//{*}Encabezado/{*}Emisor/{*}RUTEmisor")[0].text
                        except:
                            pass

                        try:
                            numTransaccion = transaccionTree.findall(".//{*}TED/{*}DD/{*}F")[0].text
                        except:
                            pass

                        try:
                            SN = transaccionTree.findall(".//{*}Referencia/{*}FolioRef")[0].text
                        except:
                            pass

                        if (IdenOrg == "" or numTransaccion == "" or SN == "" or SN[0] != "S" or len(SN) != 9 or SN[
                                                                                                                 1:9].isnumeric() == False):
                            shutil.move(pathTransaccion, dirTransaccionesOrgsinSN)

                        else:
                            carpetaTransaccion = os.path.join(dirTransaccionesOrg, numTransaccion + "_" + SN)
                            if os.path.isdir(carpetaTransaccion):
                                shutil.rmtree(carpetaTransaccion)
                            os.mkdir(carpetaTransaccion)
                            shutil.move(pathTransaccion, carpetaTransaccion)
                            nuevaDirTransaccion = os.path.join(carpetaTransaccion, arch)

                            if (carpetaTransaccion not in carpetasProcesar):
                                carpetasProcesar.append(carpetaTransaccion)

            return carpetasProcesar

        carpetasOrg = []
        carpetasTransacciones = []
        for elem in os.walk(dirIter):
            if (elem[0] == dirIter):
                for subCarp in elem[1]:
                    if (rutTest in subCarp):
                        carpetaOrg = os.path.join(dirIter, subCarp)
                        if carpetaOrg not in carpetasOrg:
                            carpetasOrg.append(carpetaOrg)
                            generarCarpetasTransaccion(carpetaOrg)

        for elem in os.walk(dirIter):
            if (elem[0] == dirIter):
                for carpOrg in elem[1]:
                    dirCarpOrg = os.path.join(dirIter, carpOrg)
                    dirCarpOrgProc = os.path.join(dirCarpOrg, "Procesar")
                    for elemProc in os.walk(dirCarpOrgProc):
                        if (elemProc[0] == dirCarpOrgProc):
                            for carpTransaccion in elemProc[1]:
                                dirCarpTransaccion = os.path.join(dirCarpOrgProc, carpTransaccion)
                                if dirCarpTransaccion not in carpetasTransacciones:
                                    carpetasTransacciones.append(dirCarpTransaccion)

        return carpetasTransacciones

    def procesarTransaccion(carpetaTransaccion):

        def getDirTransaccion(carpetaTransaccion):

            def getOrgCatalogo(RutOrg):
                import openpyxl as op

                esOrg = False
                dirCatalogoOrg = os.path.join(CurrentDirectory, "catalogos", "organizaciones.xlsx")
                wbCatalogoOrg = op.load_workbook(filename=dirCatalogoOrg)
                wsCatalogoOrg = wbCatalogoOrg.active
                lastEmptyCatalogoOrg = len(list(wsCatalogoOrg))
                k = 1
                nombreOrg = "No Identificada"
                while k < lastEmptyCatalogoOrg:
                    if (wsCatalogoOrg.cell(k, 1).value == RutOrg):
                        esOrg = True
                    k += 1
                wbCatalogoOrg.save(filename=dirCatalogoOrg)
                wbCatalogoOrg.close()

                return esOrg

            dirTransaccion = ""
            for elem in os.walk(carpetaTransaccion):
                if elem[0] == carpetaTransaccion and len(elem[2]) == 1 and ".xml" in elem[2][0] and getOrgCatalogo(elem[2][0].split("_")[0]):

                    dirTransaccion = os.path.join(carpetaTransaccion, elem[2][0])

                    if(os.path.isfile(dirTransaccion) == False):
                        dirTransaccion = ""
            return dirTransaccion

        def descargarSNAntes(dirCarpetaActual, dirTransaccion):

            from datetime import datetime
            # xml
            import lxml
            import xml.etree.ElementTree as ET
            from lxml import etree
            from lxml import etree, objectify

            # loggin
            import logging
            import glob
            import sys
            import requests
            from requests.auth import HTTPBasicAuth
            from unidecode import unidecode as ud


            def crearSolicitudSN(SN, carpetaDestino):

                dirXMLCW = ""
                if (os.path.isdir(carpetaDestino) and len(SN) == 9 and SN[0] == "S" and SN[1:].isnumeric()):
                    UniversalShipmentRequest = ET.Element("UniversalShipmentRequest", xmlns="http://www.cargowise.com/Schemas/Universal/2011/11",version="1.1")
                    ShipmentRequest = ET.SubElement(UniversalShipmentRequest, "ShipmentRequest")

                    DataContext = ET.SubElement(ShipmentRequest, "DataContext")
                    DataTargetCollection = ET.SubElement(DataContext, "DataTargetCollection")
                    DataTarget = ET.SubElement(DataTargetCollection, "DataTarget")
                    ET.SubElement(DataTarget, "Type").text = "ForwardingShipment"
                    ET.SubElement(DataTarget, "Key").text = SN
                    Company = ET.SubElement(DataContext, "Company")
                    ET.SubElement(Company, "Code").text = "SCL"
                    ET.SubElement(DataContext, "EnterpriseID").text = "A5L"
                    ET.SubElement(DataContext, "ServerID").text = "PRD"
                    # ET.SubElement(DataContext, "ServerID").text = "LIM"
                    ET.SubElement(DataContext, "DataProvider").text = "A5LSCLPRD"
                    # ET.SubElement(DataContext, "DataProvider").text = "A5LLIMPRD"

                    tree = ET.ElementTree(UniversalShipmentRequest)
                    nombreXML = datetime.today().strftime("%Y-%m-%d_%H-%M-%f") + '_ShipmentRequest_' + str(SN) + '.xml'
                    dirXMLCW = os.path.join(carpetaDestino, "Response_" + nombreXML)
                    filename = os.path.join(carpetaDestino, nombreXML)
                    tree.write(filename, encoding="utf-8", xml_declaration=False, short_empty_elements=False)
                    prettytree = lxml.etree.parse(filename)

                    with open(filename, "wb") as f:
                        f.write(etree.tostring(prettytree, pretty_print=True, xml_declaration=False, encoding="UTF-8"))
                ExtraerXML(carpetaDestino, carpetaDestino)

                return dirXMLCW

            def my_custom_logger(logger_name, level=logging.DEBUG):

                """
                Method to return a custom logger with the given name and level
                """
                logger = logging.getLogger(logger_name)
                logger.setLevel(level)
                format_string = ("%(asctime)s — %(name)s — %(levelname)s — %(funcName)s:"
                                 "%(lineno)d — %(message)s")
                log_format = logging.Formatter(format_string)
                # Creating and adding the console handler
                console_handler = logging.StreamHandler(sys.stdout)
                console_handler.setFormatter(log_format)
                logger.addHandler(console_handler)
                # Creating and adding the file handler
                file_handler = logging.FileHandler(logger_name, mode='a')
                file_handler.setFormatter(log_format)
                logger.addHandler(file_handler)
                return logger

            def ExtraerXML(InputFolderXML, OutputFolderXML):

                url = 'https://a5lprdservices.wisegrid.net/eAdaptor'

                user = "A5L.data.entry"

                dirClave = os.path.join(os.path.dirname(__file__), "clave.txt")
                password = open(dirClave).read()

                dirname = os.path.dirname(__file__)

                for file in glob.glob(os.path.join(InputFolderXML, "*.xml")):
                    try:
                        filename = os.path.basename(file)
                        data = open(file, encoding='utf-8').read()

                        response = requests.post(url, data=data.encode('utf-8'), auth=(user, password))
                        ResponseFile = open(os.path.join(OutputFolderXML, "Response_" + filename), "w+")
                        ResponseFile.write(ud(str(response.text))[4:])
                        ResponseFile.close()
                        i = 1

                        newfilename = filename
                        while os.path.exists(os.path.join(OutputFolderXML, newfilename)):
                            newfilename = filename[:-4] + "(" + str(i) + ").xml"
                            i += 1
                        direccionFila = os.path.join(InputFolderXML, filename)
                        os.remove(direccionFila)

                    except Exception as e:
                        logger = my_custom_logger(f"Response/error_log_{filename[:-4]}.log")
                        logger.debug(str(e))

            transaccionTree = ET.parse(dirTransaccion)
            SN = ""

            try:
                SN = transaccionTree.findall(".//{*}Referencia/{*}FolioRef")[0].text
            except:
                pass

            destinoShipmentCW = os.path.join(dirCarpetaActual, "respCW_" + SN)
            if os.path.isdir(destinoShipmentCW):
                shutil.rmtree(destinoShipmentCW)

            os.mkdir(destinoShipmentCW)
            return crearSolicitudSN(SN, destinoShipmentCW)

        def generarInputPostear(dirRespCW, carpetaTransaccion, dirTransaction):

            def getNombreOrgFromCode(CodeOrg):

                import openpyxl as op

                from datetime import datetime
                # xml
                import lxml
                import xml.etree.ElementTree as ET
                from lxml import etree
                from lxml import etree, objectify

                nombreOrg = ""

                dirCatalogoOrg = os.path.join(CurrentDirectory, "catalogos", "organizaciones.xlsx")
                wbCatalogoOrg = op.load_workbook(filename=dirCatalogoOrg)
                wsCatalogoOrg = wbCatalogoOrg.active

                lastEmptyCatalogoOrg = len(list(wsCatalogoOrg))
                k = 1
                encontrado = False
                while k < lastEmptyCatalogoOrg:
                    if (wsCatalogoOrg.cell(k, 3).value == CodeOrg):
                        nombreOrg = wsCatalogoOrg.cell(k, 2).value
                        encontrado = True
                    k += 1

                if (encontrado == False):
                    from datetime import datetime

                    import logging
                    import glob
                    import sys
                    import requests
                    from requests.auth import HTTPBasicAuth
                    from unidecode import unidecode as ud

                    orgLocation = os.path.join(CurrentDirectory, "catalogos", "organization")
                    if os.path.isdir(orgLocation):
                        shutil.rmtree(orgLocation)
                    os.mkdir(orgLocation)

                    def my_custom_logger(logger_name, level=logging.DEBUG):

                        """
                        Method to return a custom logger with the given name and level
                        """
                        logger = logging.getLogger(logger_name)
                        logger.setLevel(level)
                        format_string = ("%(asctime)s — %(name)s — %(levelname)s — %(funcName)s:"
                                         "%(lineno)d — %(message)s")
                        log_format = logging.Formatter(format_string)
                        # Creating and adding the console handler
                        console_handler = logging.StreamHandler(sys.stdout)
                        console_handler.setFormatter(log_format)
                        logger.addHandler(console_handler)
                        # Creating and adding the file handler
                        file_handler = logging.FileHandler(logger_name, mode='a')
                        file_handler.setFormatter(log_format)
                        logger.addHandler(file_handler)
                        return logger

                    def ExtraerXML(InputFolderXML, OutputFolderXML):

                        url = 'https://a5lprdservices.wisegrid.net/eAdaptor'

                        user = "A5L.data.entry"

                        dirClave = os.path.join(os.path.dirname(__file__), "clave.txt")
                        password = open(dirClave).read()

                        dirname = os.path.dirname(__file__)
                        responseFiles = []
                        for file in glob.glob(os.path.join(InputFolderXML, "*.xml")):
                            try:
                                filename = os.path.basename(file)
                                data = open(file, encoding='utf-8').read()

                                response = requests.post(url, data=data.encode('utf-8'), auth=(user, password))
                                ResponseFile = open(os.path.join(OutputFolderXML, "Response_" + filename), "w+")
                                ResponseFile.write(ud(str(response.text))[4:])
                                ResponseFile.close()
                                i = 1

                                newfilename = filename
                                while os.path.exists(os.path.join(OutputFolderXML, newfilename)):
                                    newfilename = filename[:-4] + "(" + str(i) + ").xml"
                                    i += 1

                                direccionFila = os.path.join(InputFolderXML, filename)
                                os.remove(direccionFila)

                            except Exception as e:
                                logger = my_custom_logger(f"Response/error_log_{filename[:-4]}.log")
                                logger.debug(str(e))
                        return responseFiles

                    NativeOrganizationRequest = ET.Element("Native", xmlns="http://www.cargowise.com/Schemas/Native")
                    BodyOrganizationRequest = ET.SubElement(NativeOrganizationRequest, "Body")

                    Organization = ET.SubElement(BodyOrganizationRequest, "Organization")
                    CriteriaGroupOrg = ET.SubElement(Organization, "CriteriaGroup", Type="Partial")
                    CriteriaEntity = ET.SubElement(CriteriaGroupOrg, "Criteria", Entity="OrgHeader",
                                                   FieldName="Code").text = CodeOrg

                    tree = ET.ElementTree(NativeOrganizationRequest)
                    filename = os.path.join(orgLocation, datetime.today().strftime(
                        "%Y %m %d_%H-%M %f") + '_OrganizationRequest_' + str(CodeOrg) + '.xml')
                    tree.write(filename, encoding="utf-8", xml_declaration=False, short_empty_elements=False)
                    prettytree = lxml.etree.parse(filename)

                    with open(filename, "wb") as f:
                        f.write(etree.tostring(prettytree, pretty_print=True, xml_declaration=False, encoding="UTF-8"))

                    responseFiles = ExtraerXML(orgLocation, orgLocation)

                    for elem in os.walk(orgLocation):
                        if elem[0] == orgLocation:
                            for arch in elem[2]:
                                if ("Response_" in arch and "_OrganizationRequest_" in arch and ".xml" in arch):
                                    dirArch = os.path.join(orgLocation, arch)
                                    treeOrg = ET.parse(dirArch)
                                    dataSourceOrgCusCodeCollection = treeOrg.findall(
                                        './/{*}Organization/{*}OrgHeader/{*}OrgCusCodeCollection')
                                    orgCorrecta = False
                                    for orgCusCode in dataSourceOrgCusCodeCollection:
                                        typeCode = orgCusCode.findall(".//{*}CodeType")[0].text
                                        countryCode = orgCusCode.findall(".//{*}CodeCountry/{*}Code")[0].text
                                        if (typeCode == "RUT" and countryCode == "CL"):
                                            orgCorrecta = True

                                    if (orgCorrecta == True):
                                        nombreOrg = treeOrg.findall('.//{*}Organization/{*}OrgHeader/{*}FullName')[
                                            0].text
                                    os.remove(dirArch)

                return nombreOrg

            def getChargeLinesPostear(carpetaTransaccion, dirTransaccion, dirRespCW):

                import lxml
                import xml.etree.ElementTree as ET
                from lxml import etree
                from lxml import etree, objectify

                treeCW = ET.parse(dirRespCW)
                treeTransaccion = ET.parse(dirTransaccion)
                creditorTransaccion = treeTransaccion.findall(".//{*}Encabezado/{*}Emisor/{*}RznSoc")[0].text
                dataSourceDetalles = treeTransaccion.findall(".//{*}Documento/{*}Detalle")

                dataSourceChargeLine = treeCW.findall(".//{*}JobCosting/{*}ChargeLineCollection/{*}ChargeLine")

                montoDetalles = 0
                montoCredCW = 0
                montoCero = False
                tuplasDetalleChLine = []
                chargeLinesPostearCostos = []
                for detalle in dataSourceDetalles:
                    detalleEnc = 0
                    chLineDetalle = []
                    montoDetTransaccion = detalle.findall(".//{*}MontoItem")[0].text

                    montoDetalles += float(montoDetTransaccion)
                    for chargeLine in dataSourceChargeLine:
                        creditorName = ""
                        creditorKey = ""
                        try:
                            creditorKey = chargeLine.findall(".//{*}Creditor/{*}Key")[0].text
                        except:
                            pass

                        if(creditorKey != ""):
                            creditorName = getNombreOrgFromCode(creditorKey)

                        if (creditorName == creditorTransaccion and creditorName != "" and
                            float(chargeLine.findall(".//{*}CostLocalAmount")[0].text) == float(montoDetTransaccion) and
                            float(chargeLine.findall(".//{*}CostLocalAmount")[0].text) > 0):
                            chLineDetalle.append(chargeLine)

                    tuplasDetalleChLine.append([detalle, chLineDetalle])

                for chargeLine in dataSourceChargeLine:

                    creditorName = ""
                    creditorKey = ""
                    try:
                        creditorKey = chargeLine.findall(".//{*}Creditor/{*}Key")[0].text

                    except:
                        pass

                    if (creditorKey != ""):
                        creditorName = getNombreOrgFromCode(creditorKey)

                    if (creditorName == creditorTransaccion and creditorName != ""):
                        montoCredCW+=float(chargeLine.findall(".//{*}CostLocalAmount")[0].text)
                        if (float(chargeLine.findall(".//{*}CostLocalAmount")[0].text) <= 0):
                            montoCero = True

                detallesEnPareja = True
                for tup in tuplasDetalleChLine:
                    if (len(tup[1]) == 1):
                        if (tup[1][0].findall(".//{*}CostIsPosted")[0].text == "false"):
                            chargeLinesPostearCostos.append(tup[1][0])

                    else:
                        detallesEnPareja = False


                if(detallesEnPareja == False or montoCero == True):
                    chargeLinesPostearCostos = []

                return chargeLinesPostearCostos, tuplasDetalleChLine, detallesEnPareja, montoCero

            def generarInputPostearEAdaptor(chargeLinesPostearCostos, montoCero):

                if (len(chargeLinesPostearCostos) > 0 and montoCero == False):
                    dirInputeAdaptor = os.path.join(CurrentDirectory, "eAdaptor", "Input")
                    dirEscturcturaCarpetasXML = os.path.join(carpetaTransaccion, "XML_CWPostCost")
                    dirInput = os.path.join(dirEscturcturaCarpetasXML, "Input")
                    dirOutput = os.path.join(dirEscturcturaCarpetasXML, "Output")

                    if os.path.isdir(dirEscturcturaCarpetasXML):
                        shutil.rmtree(dirEscturcturaCarpetasXML)

                    treeTransaccion = ET.parse(dirTransaction)
                    numTransaccion = treeTransaccion.findall(".//{*}TED/{*}DD/{*}F")[0].text

                    treeUnivShipCW = ET.parse(dirRespCW)
                    os.mkdir(dirEscturcturaCarpetasXML)
                    os.mkdir(dirInput)
                    os.mkdir(dirOutput)
                    ImportInstruction = "Update"
                    PostInstruction = "PostCost"

                    ShipmentNumber = ""
                    dataSourceDataContext = treeUnivShipCW.findall('.//{*}DataContext/{*}DataSourceCollection/{*}DataSource')
                    for dataContext in dataSourceDataContext:
                        type = dataContext.findall('.//{*}Type')[0].text
                        Key = dataContext.findall('.//{*}Key')[0].text
                        if type == "ForwardingShipment" and len(Key) == 9 and Key[0] == "S" and (Key[1:9].isnumeric()):
                            ShipmentNumber = Key

                    # ******Create XUS******
                    root = ET.Element("UniversalInterchange", xmlns="http://www.cargowise.com/Schemas/Universal/2011/11", version="1.0")
                    Header = ET.SubElement(root, "Header")
                    ET.SubElement(Header, "SenderID").text = "TESTINXML"
                    ET.SubElement(Header, "RecipientID").text = "A5LSCLPRD"

                    Body = ET.SubElement(root, "Body")

                    # ******SHIPMENT******
                    SubShipmentCollection = ET.SubElement(Body, "UniversalShipment", xmlns="http://www.cargowise.com/Schemas/Universal/2011/11", version="1.1")
                    SubShipment = ET.SubElement(SubShipmentCollection, "Shipment")

                    DataContext = ET.SubElement(SubShipment, "DataContext")
                    DataTargetCollection = ET.SubElement(DataContext, "DataTargetCollection")
                    Company = ET.SubElement(DataContext, "Company")
                    ET.SubElement(Company, "Code").text = "SCL"
                    ET.SubElement(DataContext, "EnterpriseID").text = "A5L"
                    ET.SubElement(DataContext, "ServerID").text = "PRD"
                    ET.SubElement(DataContext, "DataProvider").text = "A5LSCLPRD"

                    DataTarget = ET.SubElement(DataTargetCollection, "DataTarget")

                    ET.SubElement(DataTarget, "Type").text = "ForwardingShipment"

                    ET.SubElement(DataTarget, "Key").text = ShipmentNumber

                    # ******JobCosting******
                    JobCosting = ET.SubElement(SubShipment, "JobCosting")

                    Branch = ET.SubElement(JobCosting, "Branch")
                    ET.SubElement(Branch, "Code").text = "SCL"
                    # ET.SubElement(Branch, "Code").text = "LIM"
                    ET.SubElement(Branch, "Name").text = "AirSeaLogistics SPA"

                    Currency = ET.SubElement(JobCosting, "Currency")
                    # ET.SubElement(Currency, "Code").text = "CLP"
                    ET.SubElement(Currency, "Code").text = treeUnivShipCW.findall('.//{*}JobCosting/{*}Currency/{*}Code')[0].text

                    Department = ET.SubElement(JobCosting, "Department")
                    ET.SubElement(Department, "Code").text = treeUnivShipCW.findall('.//{*}JobCosting/{*}Department/{*}Code')[0].text

                    HomeBranch = ET.SubElement(JobCosting, "HomeBranch")
                    ET.SubElement(HomeBranch, "Code").text = "SCL"
                    # ET.SubElement(HomeBranch, "Code").text = "LIM"
                    ET.SubElement(HomeBranch, "Name").text = "AirSeaLogistics SPA"
                    # ET.SubElement(HomeBranch, "Code").text = treeUnivShipCW.findall('.//{*}JobCosting/{*}Branch/{*}Code')[0].text
                    # ET.SubElement(HomeBranch, "Name").text = treeUnivShipCW.findall('.//{*}JobCosting/{*}Branch/{*}Name')[0].text

                    # ******ChargelineCollection******
                    ChargeLineCollection = ET.SubElement(JobCosting, "ChargeLineCollection")

                    for chargeline in chargeLinesPostearCostos:
                        chargeLineChCode = chargeline.findall('.//{*}ChargeCode/{*}Code')[0].text
                        if chargeline.findall('.//{*}CostIsPosted')[0].text == "false" and float(
                                chargeline.findall('.//{*}CostOSAmount')[0].text) > 0:

                            ChargeLine = ET.SubElement(ChargeLineCollection, "ChargeLine")

                            ET.SubElement(ChargeLine, "CostAPInvoiceNumber").text = numTransaccion
                            ET.SubElement(ChargeLine, "CostInvoiceDate").text = datetime.today().strftime("%Y-%m-%d")
                            ET.SubElement(ChargeLine, "CostDueDate").text = datetime.today().strftime("%Y-%m-%d")

                            MatchingCriteriaList = []
                            MatchingCriteriaList.append(['ChargeCode', chargeLineChCode])
                            MatchingCriteriaList.append(['CostOSCurrency', chargeline.findall('.//{*}CostOSCurrency/{*}Code')[0].text])
                            MatchingCriteriaList.append(['CostOSAmount', chargeline.findall('.//{*}CostOSAmount')[0].text])
                            MatchingCriteriaList.append(['CostLocalAmount', chargeline.findall('.//{*}CostLocalAmount')[0].text])
                            MatchingCriteriaList.append(['SellOSAmount', chargeline.findall('.//{*}SellOSAmount')[0].text])
                            MatchingCriteriaList.append(['SellOSCurrency', chargeline.findall('.//{*}SellOSCurrency/{*}Code')[0].text])

                            ImportMetaData = ET.SubElement(ChargeLine, "ImportMetaData")
                            ET.SubElement(ImportMetaData, "Instruction").text = ImportInstruction
                            MatchingCriteriaCollection = ET.SubElement(ImportMetaData, "MatchingCriteriaCollection")

                            for criteria in MatchingCriteriaList:
                                MatchingCriteria = ET.SubElement(MatchingCriteriaCollection, "MatchingCriteria")
                                ET.SubElement(MatchingCriteria, "FieldName").text = criteria[0]
                                ET.SubElement(MatchingCriteria, "Value").text = criteria[1]

                            # if PostInstruction != "":
                            #     ET.SubElement(ImportMetaData, "PostingInstruction").text = PostInstruction

                    tree = ET.ElementTree(SubShipmentCollection)

                    fechaString = datetime.today().strftime("%f")

                    filename = datetime.today().strftime("%Y-%m-%d %H-%M-%S") + "_" + ShipmentNumber + "_PostearCost.xml"
                    filename = os.path.join(dirInput, filename)

                    tree.write(filename, encoding="utf-8", xml_declaration=True, short_empty_elements=False)

                    if 1 == 1:
                        prettytree = lxml.etree.parse(filename)

                        with open(filename, "wb") as f:
                            f.write(
                                etree.tostring(prettytree, pretty_print=True, xml_declaration=False, encoding="UTF-8"))

                        f.close()

                    shutil.copy(filename, dirInputeAdaptor)

            from datetime import datetime
            import lxml
            import xml.etree.ElementTree as ET
            from lxml import etree
            from lxml import etree, objectify

            chargeLinesPostearCostos, tuplasDetalleChLine, montosIguales, detallesEnPareja, montoCero = getChargeLinesPostear(carpetaTransaccion, dirTransaction, dirRespCW)

            generarInputPostearEAdaptor(chargeLinesPostearCostos, montoCero)

            return chargeLinesPostearCostos, tuplasDetalleChLine, montosIguales, detallesEnPareja, montoCero

        def ejecutarEAdaptor():
            import subprocess
            dirEjecutor = os.path.join(CurrentDirectory, "eAdaptor", "GetXUS.py")
            subprocess.call(dirEjecutor, shell=True)

        def descargarSNDespues(dirCarpetaActual, dirTransaccion):

            from datetime import datetime
            # xml
            import lxml
            import xml.etree.ElementTree as ET
            from lxml import etree
            from lxml import etree, objectify

            # loggin
            import logging
            import glob
            import sys
            import requests
            from requests.auth import HTTPBasicAuth
            from unidecode import unidecode as ud


            def crearSolicitudSN(SN, carpetaDestino):

                dirXMLCW = ""
                if (os.path.isdir(carpetaDestino) and len(SN) == 9 and SN[0] == "S" and SN[1:].isnumeric()):
                    UniversalShipmentRequest = ET.Element("UniversalShipmentRequest", xmlns="http://www.cargowise.com/Schemas/Universal/2011/11",version="1.1")
                    ShipmentRequest = ET.SubElement(UniversalShipmentRequest, "ShipmentRequest")

                    DataContext = ET.SubElement(ShipmentRequest, "DataContext")
                    DataTargetCollection = ET.SubElement(DataContext, "DataTargetCollection")
                    DataTarget = ET.SubElement(DataTargetCollection, "DataTarget")
                    ET.SubElement(DataTarget, "Type").text = "ForwardingShipment"
                    ET.SubElement(DataTarget, "Key").text = SN
                    Company = ET.SubElement(DataContext, "Company")
                    ET.SubElement(Company, "Code").text = "SCL"
                    ET.SubElement(DataContext, "EnterpriseID").text = "A5L"
                    ET.SubElement(DataContext, "ServerID").text = "PRD"
                    # ET.SubElement(DataContext, "ServerID").text = "LIM"
                    ET.SubElement(DataContext, "DataProvider").text = "A5LSCLPRD"
                    # ET.SubElement(DataContext, "DataProvider").text = "A5LLIMPRD"

                    tree = ET.ElementTree(UniversalShipmentRequest)
                    nombreXML = datetime.today().strftime("%Y-%m-%d_%H-%M-%f") + '_ShipmentRequest_' + str(SN) + '.xml'
                    dirXMLCW = os.path.join(carpetaDestino, "Response_" + nombreXML)
                    filename = os.path.join(carpetaDestino, nombreXML)
                    tree.write(filename, encoding="utf-8", xml_declaration=False, short_empty_elements=False)
                    prettytree = lxml.etree.parse(filename)

                    with open(filename, "wb") as f:
                        f.write(etree.tostring(prettytree, pretty_print=True, xml_declaration=False, encoding="UTF-8"))
                ExtraerXML(carpetaDestino, carpetaDestino)

                return dirXMLCW

            def my_custom_logger(logger_name, level=logging.DEBUG):

                """
                Method to return a custom logger with the given name and level
                """
                logger = logging.getLogger(logger_name)
                logger.setLevel(level)
                format_string = ("%(asctime)s — %(name)s — %(levelname)s — %(funcName)s:"
                                 "%(lineno)d — %(message)s")
                log_format = logging.Formatter(format_string)
                # Creating and adding the console handler
                console_handler = logging.StreamHandler(sys.stdout)
                console_handler.setFormatter(log_format)
                logger.addHandler(console_handler)
                # Creating and adding the file handler
                file_handler = logging.FileHandler(logger_name, mode='a')
                file_handler.setFormatter(log_format)
                logger.addHandler(file_handler)
                return logger

            def ExtraerXML(InputFolderXML, OutputFolderXML):

                url = 'https://a5lprdservices.wisegrid.net/eAdaptor'

                user = "A5L.data.entry"

                dirClave = os.path.join(os.path.dirname(__file__), "clave.txt")
                password = open(dirClave).read()

                dirname = os.path.dirname(__file__)

                for file in glob.glob(os.path.join(InputFolderXML, "*.xml")):
                    try:
                        filename = os.path.basename(file)
                        data = open(file, encoding='utf-8').read()

                        response = requests.post(url, data=data.encode('utf-8'), auth=(user, password))
                        ResponseFile = open(os.path.join(OutputFolderXML, "Response_" + filename), "w+")
                        ResponseFile.write(ud(str(response.text))[4:])
                        ResponseFile.close()
                        i = 1

                        newfilename = filename
                        while os.path.exists(os.path.join(OutputFolderXML, newfilename)):
                            newfilename = filename[:-4] + "(" + str(i) + ").xml"
                            i += 1
                        direccionFila = os.path.join(InputFolderXML, filename)
                        os.remove(direccionFila)

                    except Exception as e:
                        logger = my_custom_logger(f"Response/error_log_{filename[:-4]}.log")
                        logger.debug(str(e))

            transaccionTree = ET.parse(dirTransaccion)
            SN = ""

            try:
                SN = transaccionTree.findall(".//{*}Referencia/{*}FolioRef")[0].text
            except:
                pass

            destinoShipmentCW = os.path.join(dirCarpetaActual, "respCWDesp_" + SN)
            if os.path.isdir(destinoShipmentCW):
                shutil.rmtree(destinoShipmentCW)

            os.mkdir(destinoShipmentCW)
            return crearSolicitudSN(SN, destinoShipmentCW)

        def validarCambios(chargeLinesPostearCostos, tuplasDetalleChLine, dirRespCW, dirRespDespCW, dirTransaction, montosIguales, detallesEnPareja, montoCero):

            def getInfoOrg(RutOrg):

                import openpyxl as op

                from datetime import datetime
                # xml
                import lxml
                import xml.etree.ElementTree as ET
                from lxml import etree
                from lxml import etree, objectify

                codeOrg = ""
                nombreOrg = ""

                dirCatalogoOrg = os.path.join(CurrentDirectory, "catalogos", "organizaciones.xlsx")
                wbCatalogoOrg = op.load_workbook(filename=dirCatalogoOrg)
                wsCatalogoOrg = wbCatalogoOrg.active

                lastEmptyCatalogoOrg = len(list(wsCatalogoOrg))
                k = 1
                encontrado = False
                while k < lastEmptyCatalogoOrg:
                    if (wsCatalogoOrg.cell(k, 1).value == RutOrg):
                        nombreOrg = wsCatalogoOrg.cell(k, 2).value
                        codeOrg = wsCatalogoOrg.cell(k, 3).value
                        encontrado = True
                    k += 1
                if (encontrado == False):
                    from datetime import datetime

                    import logging
                    import glob
                    import sys
                    import requests
                    from requests.auth import HTTPBasicAuth
                    from unidecode import unidecode as ud

                    orgLocation = os.path.join(CurrentDirectory, "catalogos", "organization")
                    if os.path.isdir(orgLocation):
                        shutil.rmtree(orgLocation)
                    os.mkdir(orgLocation)

                    def my_custom_logger(logger_name, level=logging.DEBUG):

                        """
                        Method to return a custom logger with the given name and level
                        """
                        logger = logging.getLogger(logger_name)
                        logger.setLevel(level)
                        format_string = ("%(asctime)s — %(name)s — %(levelname)s — %(funcName)s:"
                                         "%(lineno)d — %(message)s")
                        log_format = logging.Formatter(format_string)
                        # Creating and adding the console handler
                        console_handler = logging.StreamHandler(sys.stdout)
                        console_handler.setFormatter(log_format)
                        logger.addHandler(console_handler)
                        # Creating and adding the file handler
                        file_handler = logging.FileHandler(logger_name, mode='a')
                        file_handler.setFormatter(log_format)
                        logger.addHandler(file_handler)
                        return logger

                    def ExtraerXML(InputFolderXML, OutputFolderXML):

                        url = 'https://a5lprdservices.wisegrid.net/eAdaptor'

                        user = "A5L.data.entry"

                        dirClave = os.path.join(os.path.dirname(__file__), "clave.txt")
                        password = open(dirClave).read()

                        dirname = os.path.dirname(__file__)
                        responseFiles = []
                        for file in glob.glob(os.path.join(InputFolderXML, "*.xml")):
                            try:
                                filename = os.path.basename(file)
                                data = open(file, encoding='utf-8').read()

                                response = requests.post(url, data=data.encode('utf-8'), auth=(user, password))
                                ResponseFile = open(os.path.join(OutputFolderXML, "Response_" + filename), "w+")
                                ResponseFile.write(ud(str(response.text))[4:])
                                ResponseFile.close()
                                i = 1

                                newfilename = filename
                                while os.path.exists(os.path.join(OutputFolderXML, newfilename)):
                                    newfilename = filename[:-4] + "(" + str(i) + ").xml"
                                    i += 1

                                direccionFila = os.path.join(InputFolderXML, filename)
                                os.remove(direccionFila)

                            except Exception as e:
                                logger = my_custom_logger(f"Response/error_log_{filename[:-4]}.log")
                                logger.debug(str(e))
                        return responseFiles

                    NativeOrganizationRequest = ET.Element("Native", xmlns="http://www.cargowise.com/Schemas/Native")
                    BodyOrganizationRequest = ET.SubElement(NativeOrganizationRequest, "Body")

                    Organization = ET.SubElement(BodyOrganizationRequest, "Organization")
                    CriteriaGroupOrg = ET.SubElement(Organization, "CriteriaGroup", Type="Partial")
                    CriteriaEntity = ET.SubElement(CriteriaGroupOrg, "Criteria", Entity="OrgHeader.OrgCusCode",FieldName="CustomsRegNo").text = RutOrg

                    tree = ET.ElementTree(NativeOrganizationRequest)
                    filename = os.path.join(orgLocation, datetime.today().strftime("%Y %m %d_%H-%M %f") + '_OrganizationRequest_' + str(RutOrg) + '.xml')
                    tree.write(filename, encoding="utf-8", xml_declaration=False, short_empty_elements=False)
                    prettytree = lxml.etree.parse(filename)

                    with open(filename, "wb") as f:
                        f.write(etree.tostring(prettytree, pretty_print=True, xml_declaration=False, encoding="UTF-8"))

                    responseFiles = ExtraerXML(orgLocation, orgLocation)

                    for elem in os.walk(orgLocation):
                        if elem[0] == orgLocation:
                            for arch in elem[2]:
                                if ("Response_" in arch and "_OrganizationRequest_" in arch and ".xml" in arch):
                                    dirArch = os.path.join(orgLocation, arch)
                                    treeOrg = ET.parse(dirArch)
                                    dataSourceOrgCusCodeCollection = treeOrg.findall('.//{*}Organization/{*}OrgHeader/{*}OrgCusCodeCollection')
                                    orgCorrecta = False
                                    for orgCusCode in dataSourceOrgCusCodeCollection:
                                        num = orgCusCode.findall(".//{*}CustomsRegNo")[0].text
                                        if (num == RutOrg):
                                            orgCorrecta = True

                                    if (orgCorrecta == True):
                                        codeOrg = treeOrg.findall('.//{*}Organization/{*}OrgHeader/{*}Code')[0].text
                                        nombreOrg = treeOrg.findall('.//{*}Organization/{*}OrgHeader/{*}FullName')[
                                            0].text
                                    os.remove(dirArch)
                return codeOrg, nombreOrg

            from datetime import datetime
            # xml
            import lxml
            import xml.etree.ElementTree as ET
            from lxml import etree
            from lxml import etree, objectify

            tuplasCambios = []
            cambiosValidados = []
            treeTransaction = ET.parse(dirTransaction)
            RUTCreditor = treeTransaction.findall('.//{*}Encabezado/{*}Emisor/{*}RUTEmisor')[0].text

            treeResp = ET.parse(dirRespCW)

            ShipmentNumber = ""
            dataSourceDataContext = treeResp.findall('.//{*}DataContext/{*}DataSourceCollection/{*}DataSource')
            for dataContext in dataSourceDataContext:
                type = dataContext.findall('.//{*}Type')[0].text
                Key = dataContext.findall('.//{*}Key')[0].text
                if type == "ForwardingShipment" and len(Key) == 9 and Key[0] == "S" and (Key[1:9].isnumeric()):
                    ShipmentNumber = Key

            treeRespDesp = ET.parse(dirRespDespCW)
            dataSourceChargeLinesCWDesp = treeRespDesp.findall(".//{*}ChargeLineCollection/{*}ChargeLine")
            nombreOrg, codeOrg = getInfoOrg(RUTCreditor)

            for tup in tuplasDetalleChLine:
                chLineEncontrada = []
                if(len(tup[1]) == 1):
                    chargeLinePostear = tup[1][0]
                    chargeLinePostCostLocalAmount = chargeLinePostear.findall(".//{*}CostLocalAmount")[0].text
                    chargeLinePostChargeCode = chargeLinePostear.findall(".//{*}ChargeCode/{*}Code")[0].text
                    chargeLinePostCostOSCurrency = chargeLinePostear.findall(".//{*}CostOSCurrency/{*}Code")[0].text
                    chargeLineEnc = ""

                    for chargeLineDesp in dataSourceChargeLinesCWDesp:

                        if(chargeLinePostCostLocalAmount == chargeLineDesp.findall(".//{*}CostLocalAmount")[0].text and
                        chargeLinePostChargeCode == chargeLineDesp.findall(".//{*}ChargeCode/{*}Code")[0].text and
                        chargeLinePostCostOSCurrency == chargeLineDesp.findall(".//{*}CostOSCurrency/{*}Code")[0].text):
                            chLineEncontrada.append(chargeLineDesp)

                # print("tup[1]", len(tup[1]), "chLineEncontrada", len(chLineEncontrada), "montosIguales", montosIguales, "detallesEnPareja", detallesEnPareja, "montoCero", montoCero)
                tuplasCambios.append([ShipmentNumber, nombreOrg, codeOrg, tup[0], tup[1], chLineEncontrada, montosIguales, detallesEnPareja, montoCero])

            return tuplasCambios

        dirTransaction = getDirTransaccion(carpetaTransaccion)
        tuplasCambios = []
        if(dirTransaction != ""):
            dirRespCW = descargarSNAntes(carpetaTransaccion, dirTransaction)
            chargeLinesPostearCostos, tuplasDetalleChLine, montosIguales, detallesEnPareja, montoCero = generarInputPostear(dirRespCW, carpetaTransaccion, dirTransaction)
            # ejecutarEAdaptor()
            dirRespDespCW = descargarSNDespues(carpetaTransaccion, dirTransaction)
            tuplasCambios = validarCambios(chargeLinesPostearCostos, tuplasDetalleChLine, dirRespCW, dirRespDespCW, dirTransaction, montosIguales, detallesEnPareja, montoCero)
            print("tuplasCambios", len(tuplasCambios))
        return tuplasCambios

    def generarInformeCambios(tuplasInforme, dirIter):

        def getNombreOrgFromCode(CodeOrg):

            import openpyxl as op

            from datetime import datetime
            # xml
            import lxml
            import xml.etree.ElementTree as ET
            from lxml import etree
            from lxml import etree, objectify

            nombreOrg = ""

            dirCatalogoOrg = os.path.join(CurrentDirectory, "catalogos", "organizaciones.xlsx")
            wbCatalogoOrg = op.load_workbook(filename=dirCatalogoOrg)
            wsCatalogoOrg = wbCatalogoOrg.active

            lastEmptyCatalogoOrg = len(list(wsCatalogoOrg))
            k = 1
            encontrado = False
            while k < lastEmptyCatalogoOrg:
                if (wsCatalogoOrg.cell(k, 3).value == CodeOrg):
                    nombreOrg = wsCatalogoOrg.cell(k, 2).value
                    encontrado = True
                k += 1

            if (encontrado == False):
                from datetime import datetime

                import logging
                import glob
                import sys
                import requests
                from requests.auth import HTTPBasicAuth
                from unidecode import unidecode as ud

                orgLocation = os.path.join(CurrentDirectory, "catalogos", "organization")
                if os.path.isdir(orgLocation):
                    shutil.rmtree(orgLocation)
                os.mkdir(orgLocation)

                def my_custom_logger(logger_name, level=logging.DEBUG):

                    """
                    Method to return a custom logger with the given name and level
                    """
                    logger = logging.getLogger(logger_name)
                    logger.setLevel(level)
                    format_string = ("%(asctime)s — %(name)s — %(levelname)s — %(funcName)s:"
                                     "%(lineno)d — %(message)s")
                    log_format = logging.Formatter(format_string)
                    # Creating and adding the console handler
                    console_handler = logging.StreamHandler(sys.stdout)
                    console_handler.setFormatter(log_format)
                    logger.addHandler(console_handler)
                    # Creating and adding the file handler
                    file_handler = logging.FileHandler(logger_name, mode='a')
                    file_handler.setFormatter(log_format)
                    logger.addHandler(file_handler)
                    return logger

                def ExtraerXML(InputFolderXML, OutputFolderXML):

                    url = 'https://a5lprdservices.wisegrid.net/eAdaptor'

                    user = "A5L.data.entry"

                    dirClave = os.path.join(os.path.dirname(__file__), "clave.txt")
                    password = open(dirClave).read()

                    dirname = os.path.dirname(__file__)
                    responseFiles = []
                    for file in glob.glob(os.path.join(InputFolderXML, "*.xml")):
                        try:
                            filename = os.path.basename(file)
                            data = open(file, encoding='utf-8').read()

                            response = requests.post(url, data=data.encode('utf-8'), auth=(user, password))
                            ResponseFile = open(os.path.join(OutputFolderXML, "Response_" + filename), "w+")
                            ResponseFile.write(ud(str(response.text))[4:])
                            ResponseFile.close()
                            i = 1

                            newfilename = filename
                            while os.path.exists(os.path.join(OutputFolderXML, newfilename)):
                                newfilename = filename[:-4] + "(" + str(i) + ").xml"
                                i += 1

                            direccionFila = os.path.join(InputFolderXML, filename)
                            os.remove(direccionFila)

                        except Exception as e:
                            logger = my_custom_logger(f"Response/error_log_{filename[:-4]}.log")
                            logger.debug(str(e))
                    return responseFiles

                NativeOrganizationRequest = ET.Element("Native", xmlns="http://www.cargowise.com/Schemas/Native")
                BodyOrganizationRequest = ET.SubElement(NativeOrganizationRequest, "Body")

                Organization = ET.SubElement(BodyOrganizationRequest, "Organization")
                CriteriaGroupOrg = ET.SubElement(Organization, "CriteriaGroup", Type="Partial")
                CriteriaEntity = ET.SubElement(CriteriaGroupOrg, "Criteria", Entity="OrgHeader",
                                               FieldName="Code").text = CodeOrg

                tree = ET.ElementTree(NativeOrganizationRequest)
                filename = os.path.join(orgLocation,
                                        datetime.today().strftime("%Y %m %d_%H-%M %f") + '_OrganizationRequest_' + str(
                                            CodeOrg) + '.xml')
                tree.write(filename, encoding="utf-8", xml_declaration=False, short_empty_elements=False)
                prettytree = lxml.etree.parse(filename)

                with open(filename, "wb") as f:
                    f.write(etree.tostring(prettytree, pretty_print=True, xml_declaration=False, encoding="UTF-8"))

                responseFiles = ExtraerXML(orgLocation, orgLocation)

                for elem in os.walk(orgLocation):
                    if elem[0] == orgLocation:
                        for arch in elem[2]:
                            if ("Response_" in arch and "_OrganizationRequest_" in arch and ".xml" in arch):
                                dirArch = os.path.join(orgLocation, arch)
                                treeOrg = ET.parse(dirArch)
                                dataSourceOrgCusCodeCollection = treeOrg.findall(
                                    './/{*}Organization/{*}OrgHeader/{*}OrgCusCodeCollection')
                                orgCorrecta = False
                                for orgCusCode in dataSourceOrgCusCodeCollection:
                                    typeCode = orgCusCode.findall(".//{*}CodeType")[0].text
                                    countryCode = orgCusCode.findall(".//{*}CodeCountry/{*}Code")[0].text
                                    if (typeCode == "RUT" and countryCode == "CL"):
                                        orgCorrecta = True

                                if (orgCorrecta == True):
                                    nombreOrg = treeOrg.findall('.//{*}Organization/{*}OrgHeader/{*}FullName')[0].text
                                os.remove(dirArch)

            return nombreOrg

        import openpyxl as op

        from datetime import datetime
        # xml
        import lxml
        import xml.etree.ElementTree as ET
        from lxml import etree
        from lxml import etree, objectify

        nuevaDir = os.path.join(dirIter, "plantillaReporteAcciones.xlsx")
        if(os.path.isfile(nuevaDir)):
            os.remove(nuevaDir)

        dirPlantillaInforme = os.path.join(CurrentDirectory, "catalogos", "plantillaReporteAcciones.xlsx")
        shutil.copy(dirPlantillaInforme, dirIter)
        nuevaDir = os.path.join(dirIter, "plantillaReporteAcciones.xlsx")
        wbInformeCambios = op.load_workbook(filename=nuevaDir)
        wsInformeCambios = wbInformeCambios.active
        # ShipmentNumber, nombreOrg, codeOrg, detalle, chLineAntes, chLineDespues, montosIguales, detallesEnPareja, montoCero

        arrInfoBase = []
        for tup in tuplasInforme:
            if([tup[0], tup[1], tup[2]] not in arrInfoBase):
                arrInfoBase.append([tup[0], tup[1], tup[2]])

        for infoBase in arrInfoBase:
            for tup in tuplasInforme:
                if (tup[0] == infoBase[0] and tup[1] == infoBase[1] and tup[2] == infoBase[2]):
                    # print("____________________________________________________________________________")
                    # print("SN", tup[0])
                    # print("nombreOrg", tup[1])
                    # print("codeOrg", tup[2])
                    # print("MontoDetalle", tup[3].findall(".//{*}MontoItem")[0].text)
                    # print("chLineAntes", len(tup[4]))
                    # print("chLineDespues", len(tup[5]))
                    # print("montosIguales", tup[6])
                    # print("detallesEnPareja", tup[7])
                    # print("montoCero", tup[8])
                    esc = len(list(wsInformeCambios))+1
                    wsInformeCambios.cell(esc, 1).value = tup[0]
                    # F = tup[3].findall(".//{*}TED/{*}DD/{*}F")[0].text
                    # print("F", F)
                    # wsInformeCambios.cell(esc, 2).value = tup[3].findall(".//{*}TED/{*}DD/{*}F")[0].text
                    wsInformeCambios.cell(esc, 3).value = tup[2]+"("+tup[1]+")"
                    wsInformeCambios.cell(esc, 4).value = tup[3].findall(".//{*}MontoItem")[0].text
                    if(tup[6] and tup[8] == False):
                        if(len(tup[4]) == 1):
                            if(tup[4][0].findall(".//{*}CostIsPosted")[0].text == "false"):
                                nombreOrg = getNombreOrgFromCode(tup[4][0].findall(".//{*}Creditor/{*}Key")[0].text)
                                wsInformeCambios.cell(esc, 6).value = nombreOrg
                                wsInformeCambios.cell(esc, 7).value = tup[4][0].findall(".//{*}CostLocalAmount")[0].text
                                wsInformeCambios.cell(esc, 8).value = "Un Billing coincide"
                                wsInformeCambios.cell(esc, 9).value = "Se debe postear el billing"
                                if (len(tup[5]) == 1):
                                    CostIsPosted = tup[6][0].findall(".//{*}CostIsPosted")[0].text
                                    if(CostIsPosted == "false"):
                                        wsInformeCambios.cell(esc, 8).value = "No se posteó el billing"

                                    if(CostIsPosted == "true"):
                                        wsInformeCambios.cell(esc, 8).value = "Se posteó el billing"

                            elif(tup[4][0].findall(".//{*}CostIsPosted")[0].text == "true"):
                                nombreOrg = getNombreOrgFromCode(tup[4][0].findall(".//{*}Creditor/{*}Key")[0].text)
                                wsInformeCambios.cell(esc, 5).value = nombreOrg
                                nombreOrg = getNombreOrgFromCode(tup[4][0].findall(".//{*}Creditor/{*}Key")[0].text)
                                wsInformeCambios.cell(esc, 6).value = nombreOrg
                                wsInformeCambios.cell(esc, 7).value = tup[4][0].findall(".//{*}CostLocalAmount")[0].text
                                wsInformeCambios.cell(esc, 8).value = "Un billing coincide, pero ya había sido posteado"
                                wsInformeCambios.cell(esc, 9).value = "-"
                                wsInformeCambios.cell(esc, 10).value = "No se realizó ningún cambio en CW"

                        else:
                            wsInformeCambios.cell(esc, 5).value = "-"
                            if(len(tup[4]) > 1):
                                wsInformeCambios.cell(esc, 6).value = "Más de un Billing coincide"

                            elif(len(tup[4]) == 0):
                                wsInformeCambios.cell(esc, 6).value = "Ningún Billing coincide"

                            wsInformeCambios.cell(esc, 7).value = "-"
                            wsInformeCambios.cell(esc, 8).value = "No se realizó ningún cambio en CW"
                else:
                    wsInformeCambios.cell(esc, 4).value = "-"
                    wsInformeCambios.cell(esc, 5).value = "-"
                    conclusion = ""
                    if(tup[6] == False):
                        conclusion+="- El monto del Creditor no coincide con la factura "

                    if(tup[8] == True):
                        conclusion+="- El Creditor tenía Billings de costo 0 en el shipment en CW "

                    wsInformeCambios.cell(esc, 6).value = conclusion
                    wsInformeCambios.cell(esc, 7).value = "-"
                    wsInformeCambios.cell(esc, 8).value = "No se realizó ningún cambio en CW"

        wbInformeCambios.save(filename=nuevaDir)
        wbInformeCambios.close()

    tuplasCambios = []
    carpetasTransacciones = getCarpetasTransacciones(dirIter)

    for carpetaTransaccion in carpetasTransacciones:
        print("carpetaTransaccion", carpetaTransaccion)
        cambiosValidados = procesarTransaccion(carpetaTransaccion)
        for tup in cambiosValidados:
            tuplasCambios.append(tup)


    generarInformeCambios(tuplasCambios, dirIter)

# getDescItem()

# actualizarCatalogoOrgs()
# dirIter = generarEstructuraIteracion()
dirIter = os.path.join(CurrentDirectory, "procesando", "iteracion_2022-07-14")
procesarEstructuraIteracion(dirIter)

# asegurarEstructuraTransacciones()
# getCodesOrg()